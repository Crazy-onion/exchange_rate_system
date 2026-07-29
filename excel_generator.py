"""
Excel 模板生成模块
复制模板文件，更新数据Sheet，保持公式Sheet不变
支持任意目标月份生成（含sheet重命名+公式引用更新）
"""
import shutil
import openpyxl
from datetime import datetime, timedelta
from copy import copy
import holidays


def date_to_excel_serial(d):
    """日期转Excel序列号 (Excel 1900日期系统)"""
    if isinstance(d, str):
        d = datetime.strptime(d, '%Y-%m-%d').date()
    return d.toordinal() - 730349  # 1900-01-01 = 1, accounting for 1900 leap year bug


def generate_excel(template_path, output_path, ortax_cny, ortax_usd, rmb_rates, rmb_headers, converter_data, target_year, target_month):
    """
    生成汇率底稿Excel (兼容旧接口)
    """
    return generate_excel_for_month(
        template_path, output_path,
        ortax_cny, ortax_usd, rmb_rates, rmb_headers, converter_data,
        target_year, target_month
    )


def generate_excel_for_month(template_path, output_path, ortax_cny, ortax_usd,
                              rmb_rates, rmb_headers, converter_data,
                              target_year, target_month):
    """
    为指定月份生成汇率底稿Excel
    - 复制模板
    - 重命名Sheet匹配目标月份
    - 更新所有公式中的Sheet引用
    - 填入正确月份的数据
    """
    print(f"\n[Excel] 生成{target_year}年{target_month}月汇率底稿: {output_path}")

    # 1. 复制模板
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    # 2. 计算Sheet重命名映射
    prev_month = target_month - 1 if target_month > 1 else 12
    prev_year = target_year if target_month > 1 else target_year - 1

    sheet_renames = {}
    for name in wb.sheetnames:
        if '人民币汇率中间价-' in name and '月初' not in name:
            new_name = f"人民币汇率中间价-{target_year % 100:02d}年{target_month:02d}月"
            if name != new_name:
                sheet_renames[name] = new_name
        elif '印尼央行汇率-' in name:
            new_name = f"印尼央行汇率-{target_month}月"
            if name != new_name:
                sheet_renames[name] = new_name
        elif '各种货币对美元折算率-' in name:
            new_name = f"各种货币对美元折算率-{prev_month}月"
            if name != new_name:
                sheet_renames[name] = new_name

    # 3. 重命名Sheet
    for old_name, new_name in sheet_renames.items():
        wb[old_name].title = new_name
        print(f"  Sheet重命名: {old_name} -> {new_name}")

    # 4. 更新所有Sheet中的公式引用 (替换旧Sheet名为新Sheet名)
    if sheet_renames:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        for old_name, new_name in sheet_renames.items():
                            # 替换带引号的引用: '旧名'! -> '新名'!
                            cell.value = cell.value.replace(f"'{old_name}'!", f"'{new_name}'!")
                            # 替换不带引号的引用: 旧名! -> 新名!
                            cell.value = cell.value.replace(f"{old_name}!", f"{new_name}!")

    # 5. 更新数据Sheet
    _update_rmb_rate_sheet(wb, rmb_rates, rmb_headers, target_year, target_month)
    _update_rmb_summary_sheet(wb, rmb_rates, rmb_headers, target_year, target_month)
    _update_converter_sheet_for_month(wb, converter_data, target_year, target_month)
    _update_ortax_sheet(wb, ortax_cny, ortax_usd, target_year, target_month)
    _update_draft_sheet(wb, target_year, target_month)

    # 6. 保存
    wb.save(output_path)
    print(f"[Excel] 完成: {output_path}")
    return output_path


def _update_rmb_rate_sheet(wb, rmb_rates, rmb_headers, year, month):
    """更新人民币汇率中间价-{年}{月} Sheet"""
    # 查找匹配的sheet (按模式匹配，不依赖具体月份名)
    target_sheet = None
    for name in wb.sheetnames:
        if '人民币汇率中间价-' in name and '月初' not in name:
            target_sheet = name

    if not target_sheet:
        print(f"  [人民币汇率中间价] 未找到Sheet")
        return

    ws = wb[target_sheet]
    print(f"  更新Sheet: {target_sheet}")

    # 写表头
    for col_idx, header in enumerate(rmb_headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写数据 (按日期升序)
    month_rates = [r for r in rmb_rates if r['date'].startswith(f"{year}-{month:02d}-")]
    month_rates.sort(key=lambda x: x['date'])

    for row_idx, rate in enumerate(month_rates, 2):
        ws.cell(row=row_idx, column=1, value=rate['date'])
        for col_idx, header in enumerate(rmb_headers[1:], 2):
            if header in rate:
                ws.cell(row=row_idx, column=col_idx, value=rate[header])

    print(f"    写入{len(month_rates)}条记录")


def _update_rmb_summary_sheet(wb, rmb_rates, rmb_headers, year, month):
    """更新人民币汇率中间价（月初，月末）Sheet"""
    sheet_name = "人民币汇率中间价（月初，月末）"
    if sheet_name not in wb.sheetnames:
        print(f"  [月初月末] 未找到Sheet")
        return

    ws = wb[sheet_name]
    print(f"  更新Sheet: {sheet_name}")

    # 获取目标月的数据 (rmb_rates 是降序的)
    month_rates = [r for r in rmb_rates if r['date'].startswith(f"{year}-{month:02d}-")]

    if not month_rates:
        print(f"    无{year}-{month:02d}数据")
        return

    # 月初: 当月第一个工作日的汇率
    # 如果1号是假期，取上个月最后一个有汇率的工作日
    first_day = datetime(year, month, 1).date()
    if holidays.is_holiday(first_day):
        # 取上个月最后一个有汇率的工作日
        prev_rates = [r for r in rmb_rates if r['date'] < f"{year}-{month:02d}-01"]
        if prev_rates:
            month_start = prev_rates[0]  # 降序排列，第一条是最近的
        else:
            month_start = month_rates[-1]  # fallback
    else:
        # 取当月第一条数据 (降序排列中最后一条 = 最早日期)
        month_start = month_rates[-1]

    # 月末: 当月最新汇率 (降序排列中第一条 = 最新日期)
    month_end = month_rates[0]

    # 写月初 (Row 2)
    ws.cell(row=2, column=1, value=month_start['date'])
    for col_idx, header in enumerate(rmb_headers[1:], 2):
        if header in month_start:
            ws.cell(row=2, column=col_idx, value=month_start[header])

    # 写月末 (Row 3)
    ws.cell(row=3, column=1, value=month_end['date'])
    for col_idx, header in enumerate(rmb_headers[1:], 2):
        if header in month_end:
            ws.cell(row=3, column=col_idx, value=month_end[header])

    print(f"    月初: {month_start['date']}, 月末: {month_end['date']}")


def _update_converter_sheet_for_month(wb, converter_data, year, month):
    """更新各种货币对美元折算率 Sheet - 取目标月上个月末的数据"""
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1

    if not converter_data:
        print(f"  [折算率] 无数据")
        return

    # 找到上个月末的折算率数据
    # converter_data 的 key 是日期字符串如 '2026-06-30'
    target_prefix = f"{prev_year}-{prev_month:02d}-"
    matching_dates = [d for d in converter_data.keys() if d.startswith(target_prefix)]

    if not matching_dates:
        # 回退：找目标月1号之前最近的一期
        target_start = f"{year}-{month:02d}-01"
        matching_dates = [d for d in converter_data.keys() if d < target_start]

    if not matching_dates:
        print(f"  [折算率] 无{prev_year}-{prev_month:02d}数据")
        return

    # 取最新的一期
    latest_date = sorted(matching_dates, reverse=True)[0]
    rates = converter_data[latest_date]

    # 找到折算率Sheet
    target_sheet = None
    for name in wb.sheetnames:
        if '各种货币对美元折算率' in name:
            target_sheet = name

    if not target_sheet:
        print(f"  [折算率] 未找到Sheet")
        return

    ws = wb[target_sheet]
    print(f"  更新Sheet: {target_sheet}")

    # 更新日期
    date_obj = datetime.strptime(latest_date, '%Y-%m-%d')
    ws.cell(row=3, column=2, value=f"（{date_obj.year}年{date_obj.month}月{date_obj.day}日）")

    # 更新折算率数据 (两列结构: B/C/D/E 和 F/G/H/I)
    for row in ws.iter_rows(min_row=7, max_row=56):
        for col_start in [2, 6]:  # B列和F列
            code_cell = row[col_start - 1]
            if code_cell.value and len(str(code_cell.value).strip()) == 3:
                code = str(code_cell.value).strip()
                if code in rates:
                    rate_cell = row[col_start + 2]  # E列或I列
                    rate_cell.value = rates[code]

    print(f"    数据日期: {latest_date}, 货币数: {len(rates)}")


def _get_ortax_month_start(rates, year, month):
    """
    获取月初汇率: 当月第一个工作日
    如果1号是假期，取上个月最后一个工作日
    rates: 降序排列
    """
    first_day = datetime(year, month, 1).date()
    if holidays.is_holiday(first_day):
        # 取上个月最后一个有汇率的工作日
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        prev_rates = [r for r in rates if r['date'].startswith(f"{prev_year}-{prev_month:02d}-")]
        # 降序排列，第一条是最新 = 上月最后工作日
        return prev_rates[0] if prev_rates else None
    else:
        # 取当月最早日期 (降序中最后一条)
        month_rates = [r for r in rates if r['date'].startswith(f"{year}-{month:02d}-")]
        return month_rates[-1] if month_rates else None


def _get_ortax_month_end(rates, year, month):
    """
    获取月末汇率: 当月最后一个工作日
    rates: 降序排列，取第一条匹配 = 最新日期
    """
    month_rates = [r for r in rates if r['date'].startswith(f"{year}-{month:02d}-")]
    return month_rates[0] if month_rates else None


def _update_ortax_sheet(wb, ortax_cny, ortax_usd, year, month):
    """更新印尼央行汇率 Sheet"""
    target_sheet = None
    for name in wb.sheetnames:
        if '印尼央行汇率' in name:
            target_sheet = name

    if not target_sheet:
        print(f"  [印尼央行] 未找到Sheet")
        return

    ws = wb[target_sheet]
    print(f"  更新Sheet: {target_sheet}")

    # 月初 = 当月第一个工作日 (或上月最后工作日如果1号是假期)
    cny_start = _get_ortax_month_start(ortax_cny, year, month)
    usd_start = _get_ortax_month_start(ortax_usd, year, month)

    # 月末 = 上月最后工作日
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    cny_end = _get_ortax_month_end(ortax_cny, prev_year, prev_month)
    usd_end = _get_ortax_month_end(ortax_usd, prev_year, prev_month)

    if not cny_start and not cny_end:
        print(f"    无{year}-{month:02d}印尼央行数据")
        return

    # 找到模板中最后两个数据区块
    last_date_row = None
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            last_date_row = row
            break

    if not last_date_row:
        print(f"    未找到数据区块")
        return

    # 最后一个区块 = 月初 (当月第一个工作日)
    if cny_start:
        ws.cell(row=last_date_row, column=1, value=datetime.strptime(cny_start['date'], '%Y-%m-%d'))
        ws.cell(row=last_date_row + 2, column=15, value=cny_start['rate'])
        ws.cell(row=last_date_row + 3, column=15, value=f"=1/O{last_date_row + 2}")
    if usd_start:
        ws.cell(row=last_date_row + 6, column=15, value=usd_start['rate'])
        ws.cell(row=last_date_row + 7, column=15, value=f"=1/O{last_date_row + 6}")

    # 倒数第二个区块 = 月末 (上月最后工作日)
    prev_date_row = None
    for row in range(last_date_row - 1, max(1, last_date_row - 15), -1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            prev_date_row = row
            break

    if prev_date_row:
        if cny_end:
            ws.cell(row=prev_date_row, column=1, value=datetime.strptime(cny_end['date'], '%Y-%m-%d'))
            ws.cell(row=prev_date_row + 2, column=15, value=cny_end['rate'])
            ws.cell(row=prev_date_row + 3, column=15, value=f"=1/O{prev_date_row + 2}")
        if usd_end:
            ws.cell(row=prev_date_row + 6, column=15, value=usd_end['rate'])
            ws.cell(row=prev_date_row + 7, column=15, value=f"=1/O{prev_date_row + 6}")

    print(f"    月末({prev_date_row}行): CNY={cny_end['rate'] if cny_end else 'N/A'}, USD={usd_end['rate'] if usd_end else 'N/A'}")
    print(f"    月初({last_date_row}行): CNY={cny_start['rate'] if cny_start else 'N/A'}, USD={usd_start['rate'] if usd_start else 'N/A'}")


def _update_draft_sheet(wb, year, month):
    """更新底稿Sheet的日期"""
    if '底稿' not in wb.sheetnames:
        return

    ws = wb['底稿']
    print(f"  更新底稿日期")

    # 计算月末日期 (上月最后一天)
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    if prev_month == 12:
        last_day = datetime(prev_year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(prev_year, prev_month + 1, 1) - timedelta(days=1)

    date_str = f"{prev_year}.{prev_month:02d}.{last_day.day:02d}"

    # 更新B列日期 (B2-B34)
    for row in range(2, 35):
        ws.cell(row=row, column=2, value=date_str)

    print(f"    底稿日期: {date_str}")
